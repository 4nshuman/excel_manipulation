import openpyxl
from os.path import join, dirname, abspath, realpath
import datetime, get_file_data as gfd

def get_sheet(sheet_name):
    global xfile
    xfile = openpyxl.load_workbook(src_file,  read_only=False, keep_vba=True)
    return (xfile.get_sheet_by_name(sheet_name))

def sheet_write_func(readable_sheet, write_sheet_name):
    if (write_sheet_name=='Total Portfolio-slide 12'):
         business_impact = data_dict['total_time']
         p1_p2_incidents = data_dict['total_p1_p2']
         total_incidents = data_dict['total_tickets']
    else:
        business_impact = data_dict[write_sheet_name.lower()+'_time']
        p1_p2_incidents = data_dict[write_sheet_name.lower()+'_p1_p2_count']
        total_incidents = data_dict[write_sheet_name.lower()+'_total_count']

    sheet = get_sheet(write_sheet_name)
    start_date = datetime.datetime(1899,12,30)
    current_date = datetime.datetime.now()
    lastmonth = current_date.month-1 if current_date.month>1 else 12
    lastyear = current_date.year-1 if lastmonth==12 else current_date.year

    prev_date = datetime.datetime(lastyear, lastmonth, 1)
    current_date = datetime.datetime(current_date.year, current_date.month, 1)

    # Print all values, iterating through rows and columns
    num_cols = readable_sheet.ncols   # Number of columns
    for row_idx in range(0,readable_sheet.nrows):
        for col_idx in range(0, num_cols):  # Iterate through columns
            cell_obj = readable_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
            #print (cell_obj.value)
            if(cell_obj.value == (prev_date-start_date).days):
                mon_col_idx=col_idx
                mon_row_idx=row_idx

    for row in sheet.iter_rows(min_col=3, min_row=26, max_col=3, max_row=32):
        for cell in row:
            #print(row)
            if cell.row==26 :
                sheet.cell(row=cell.row, column=mon_col_idx+2).value = (current_date-start_date).days
            if cell.row==27 :
                sheet.cell(row=cell.row, column=mon_col_idx+2).value = business_impact
            if cell.row==28 :
                sheet.cell(row=cell.row, column=mon_col_idx+2).value = p1_p2_incidents
            if cell.row==29 :
                sheet.cell(row=cell.row, column=mon_col_idx+2).value = 'fill data'
            if cell.row==30 :
                sheet.cell(row=cell.row, column=mon_col_idx+2).value = 'fill data'
            if cell.row==31 :
                sheet.cell(row=cell.row, column=mon_col_idx+2).value = 'fill data'
            if cell.row==32 :
                sheet.cell(row=cell.row, column=mon_col_idx+2).value = total_incidents
            if cell.has_style:
                sheet.cell(row=cell.row, column=mon_col_idx+2).style = cell.style
    xfile.save(src_file)
    global msg
    msg="Incidient data for sheet : "+write_sheet_name+" written successfully"
    print(msg)
    return msg

def write_total_incidents(readable_sheet, write_sheet_name):
    sheet = get_sheet(write_sheet_name)
    num_cols = readable_sheet.ncols   # Number of columns
    flg=0
    for row_idx in range(0,readable_sheet.nrows):
        for col_idx in range(0, num_cols):  # Iterate through columns
            cell_obj = readable_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
            #print (cell_obj.value)
            if(cell_obj.value == 'Reported'):
                reported_col_idx=col_idx+1
            elif(cell_obj.value == 'Resolved'):
                resolved_col_idx=col_idx+1
            elif(cell_obj.value == 'Closed'):
                closed_col_idx=col_idx+1
            elif(cell_obj.value == 'Open Status'):
                open_col_idx=col_idx+1
                heading_row = row_idx+1
                flg=1 # Finished fetching all columns needed so now need to break
        if(flg): # if all columns fetched then break
            break


    for row in sheet.iter_rows(min_row=heading_row, max_row=heading_row+7):
        for cell in row:
            if cell.row==heading_row+1 :
                sheet.cell(row=cell.row, column=reported_col_idx).value = data_dict['total_tickets']
            elif cell.row==heading_row+2 :
                sheet.cell(row=cell.row, column=reported_col_idx).value = data_dict['ecc_total_count']
                sheet.cell(row=cell.row, column=resolved_col_idx).value = data_dict['ecc_resolved']
                sheet.cell(row=cell.row, column=closed_col_idx).value = data_dict['ecc_closed']
                sheet.cell(row=cell.row, column=open_col_idx).value = data_dict['ecc_open']
            elif cell.row==heading_row+3 :
                sheet.cell(row=cell.row, column=reported_col_idx).value = data_dict['filenet_total_count']
                sheet.cell(row=cell.row, column=resolved_col_idx).value = data_dict['filenet_resolved']
                sheet.cell(row=cell.row, column=closed_col_idx).value = data_dict['filenet_closed']
                sheet.cell(row=cell.row, column=open_col_idx).value = data_dict['filenet_open']
            elif cell.row==heading_row+4 :
                sheet.cell(row=cell.row, column=reported_col_idx).value = data_dict['seds_total_count']
                sheet.cell(row=cell.row, column=resolved_col_idx).value = data_dict['seds_resolved']
                sheet.cell(row=cell.row, column=closed_col_idx).value = data_dict['seds_closed']
                sheet.cell(row=cell.row, column=open_col_idx).value = data_dict['seds_open']
            elif cell.row==heading_row+5 :
                sheet.cell(row=cell.row, column=reported_col_idx).value = data_dict['kofax_total_count']
                sheet.cell(row=cell.row, column=resolved_col_idx).value = data_dict['kofax_resolved']
                sheet.cell(row=cell.row, column=closed_col_idx).value = data_dict['kofax_closed']
                sheet.cell(row=cell.row, column=open_col_idx).value = data_dict['kofax_open']
            elif cell.row==heading_row+6 :
                sheet.cell(row=cell.row, column=reported_col_idx).value = data_dict['ilinx_total_count']
                sheet.cell(row=cell.row, column=resolved_col_idx).value = data_dict['ilinx_resolved']
                sheet.cell(row=cell.row, column=closed_col_idx).value = data_dict['ilinx_closed']
                sheet.cell(row=cell.row, column=open_col_idx).value = data_dict['ilinx_open']
            elif cell.row==heading_row+7 :
                sheet.cell(row=cell.row, column=reported_col_idx).value = data_dict['alexandria_total_count']
                sheet.cell(row=cell.row, column=resolved_col_idx).value = data_dict['alexandria_resolved']
                sheet.cell(row=cell.row, column=closed_col_idx).value = data_dict['alexandria_closed']
                sheet.cell(row=cell.row, column=open_col_idx).value = data_dict['alexandria_open']

    xfile.save(src_file)
    global msg
    msg="Incident data for sheet : "+write_sheet_name+" written successfully"
    print(msg)
    return msg

def incident_write(src_f, dir_name,incident_data_dict):
    global data_dict,src_file
    src_file = src_f
    data_dict = incident_data_dict

    import logging
    LOG_FILENAME = 'SCRIPT.log'
    logging.basicConfig(filename=LOG_FILENAME,level=logging.INFO)
    file_dest_path = join(dirname(dirname(abspath(__file__))), dir_name, src_f)

    sheet_write_func(gfd.get_sheet_readable(src_f, 4, dir_name), 'Total Portfolio-slide 12')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    sheet_write_func(gfd.get_sheet_readable(src_f, 5, dir_name), 'ECC')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    sheet_write_func(gfd.get_sheet_readable(src_f, 6, dir_name), 'FileNet')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    sheet_write_func(gfd.get_sheet_readable(src_f, 7, dir_name), 'SEDS')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    sheet_write_func(gfd.get_sheet_readable(src_f, 8, dir_name), 'Kofax')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    sheet_write_func(gfd.get_sheet_readable(src_f, 9, dir_name), 'ILINX')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    sheet_write_func(gfd.get_sheet_readable(src_f, 10, dir_name), 'Alexandria')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    write_total_incidents(gfd.get_sheet_readable(src_f, 14, dir_name), 'Total Incidents')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
