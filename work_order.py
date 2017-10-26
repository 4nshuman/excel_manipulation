from __future__ import print_function
from os.path import join, dirname, abspath
import xlrd
import get_file_data as gfd

# Vars initializations --start
total_count_ecc=comp_count_ecc=total_count_filenet=comp_count_filenet=0
total_count_seds=comp_count_seds=total_count_kofax=comp_count_kofax=0
total_count_ilinx=comp_count_ilinx=total_count_alexandria=comp_count_alexandria=0
time_ecc=time_filenet=time_seds=time_kofax=time_ilinx=time_alexandria=0
# Vars initializations --end

def generate_column_indices(xl_sheet):
    global app_col_idx, time_col_idx, status_col_idx
    num_cols = xl_sheet.ncols   # Number of columns
    for col_idx in range(0, num_cols):  # Iterate through columns
        cell_obj = xl_sheet.cell(0, col_idx)  # Get cell object by row, col
        if(cell_obj.value == 'App Name'):
            #print("Found at "+str(col_idx))
            app_col_idx=col_idx
        elif(cell_obj.value == 'Time'):
            #print("Found at "+str(col_idx))
            time_col_idx=col_idx
        elif(cell_obj.value == 'Status'):
            #print("Found at "+str(col_idx))
            status_col_idx=col_idx

def calculations(xl_sheet):
    # GETTING APPLICATION COUNT
    global total_count_ecc, comp_count_ecc, total_count_filenet, comp_count_filenet, total_count_seds, comp_count_seds
    global total_count_kofax, comp_count_kofax, total_count_ilinx, comp_count_ilinx, total_count_alexandria, comp_count_alexandria
    for row_idx in range(0, xl_sheet.nrows):    # Iterate through rows
        cell_obj = xl_sheet.cell(row_idx, app_col_idx)  # Get cell object by row, col
        status_cell_obj = xl_sheet.cell(row_idx, status_col_idx)  # Get cell object by row, col
        if(cell_obj.value=='ECC'):
            if(status_cell_obj.value=='COMP'):
                comp_count_ecc+=1
            total_count_ecc+=1
        elif(cell_obj.value=='FileNet'):
            if(status_cell_obj.value=='COMP'):
                comp_count_filenet+=1
            total_count_filenet+=1
        elif(cell_obj.value=='SEDS'):
            if(status_cell_obj.value=='COMP'):
                comp_count_seds+=1
            total_count_seds+=1
        elif(cell_obj.value=='Kofax'):
            if(status_cell_obj.value=='COMP'):
                comp_count_kofax+=1
            total_count_kofax+=1
        elif(cell_obj.value=='ILINX'):
            if(status_cell_obj.value=='COMP'):
                comp_count_ilinx+=1
            total_count_ilinx+=1
        elif(cell_obj.value=='Alexandria'):
            if(status_cell_obj.value=='COMP'):
                comp_count_alexandria+=1
            total_count_alexandria+=1
        #print ('Column: [%s] cell_value: [%s]' % (col_idx, cell_obj.value))

    # GETTING APPLICATION AVERAGE TIME TAKEN
    global time_ecc, time_filenet, time_seds, time_kofax, time_ilinx, time_alexandria
    for row_idx in range(1, xl_sheet.nrows):    # Iterate through rows
        cell_obj = xl_sheet.cell(row_idx, app_col_idx)  # Get cell object by row, col
        cell_time_obj=xl_sheet.cell(row_idx, time_col_idx)
        if(cell_time_obj.value == ''):
            continue;
        elif(cell_obj.value=='ECC'):
            time_ecc+=float(cell_time_obj.value)
        elif(cell_obj.value=='FileNet'):
            time_filenet+=float(cell_time_obj.value)
        elif(cell_obj.value=='SEDS'):
            time_seds+=float(cell_time_obj.value)
        elif(cell_obj.value=='Kofax'):
            time_kofax+=float(cell_time_obj.value)
        elif(cell_obj.value=='ILINX'):
            time_ilinx+=float(cell_time_obj.value)
        elif(cell_obj.value=='Alexandria'):
            time_alexandria+=float(cell_time_obj.value)
        #print ('Column: [%s] cell_value: [%s]' % (col_idx, cell_obj.value))

def display():
    global total_count_ecc, comp_count_ecc, total_count_filenet, comp_count_filenet, total_count_seds, comp_count_seds
    global total_count_kofax, comp_count_kofax, total_count_ilinx, comp_count_ilinx, total_count_alexandria, comp_count_alexandria
    global time_ecc, time_filenet, time_seds, time_kofax, time_ilinx, time_alexandria
    print("########### PRINTING TIME DATA ###########")
    print('ECC: [%s]' % ((time_ecc/comp_count_ecc) if comp_count_ecc !=0 else 0))
    print('FileNet: [%s]' % ((time_filenet/comp_count_filenet) if comp_count_filenet !=0 else 0))
    print('SEDS: [%s]' % ((time_seds/comp_count_seds) if comp_count_seds !=0 else 0))
    print('Kofax: [%s]' % ((time_kofax/comp_count_kofax) if comp_count_kofax !=0 else 0))
    print('ILINX: [%s]' % ((time_ilinx/comp_count_ilinx) if comp_count_ilinx !=0 else 0))
    print('Alexandria: [%s]' % ((time_alexandria/comp_count_alexandria) if comp_count_alexandria !=0 else 0))
    print("########### PRINTING COUNT DATA ###########")
    print('ECC: total [%s] complete [%s]' %(total_count_ecc,comp_count_ecc))
    print('FileNet: total [%s] complete [%s]' %(total_count_filenet,comp_count_filenet))
    print('SEDS: total [%s] complete [%s]' %(total_count_seds,comp_count_seds))
    print('Kofax: total [%s] complete [%s]' %(total_count_kofax,comp_count_kofax))
    print('ILINX: total [%s] complete [%s]' %(total_count_ilinx,comp_count_ilinx))
    print('Alexandria: total [%s] complete [%s]' %(total_count_alexandria,comp_count_alexandria))

def write_file(src_filename):
    import openpyxl
    xfile = openpyxl.load_workbook(src_filename,  read_only=False, keep_vba=True)
    sheet=xfile.get_sheet_by_name('WO Data')
    #WRITING TOTAL TICKETS
    sheet['B6'] = total_count_ecc
    sheet['B7'] = total_count_filenet
    sheet['B8'] = total_count_seds
    sheet['B9'] = total_count_kofax
    sheet['B10'] = total_count_ilinx
    sheet['B11'] = total_count_alexandria
    #WRITING COMPLETED TICKETS
    sheet['C6'] = comp_count_ecc
    sheet['C7'] = comp_count_filenet
    sheet['C8'] = comp_count_seds
    sheet['C9'] = comp_count_kofax
    sheet['C10'] = comp_count_ilinx
    sheet['C11'] = comp_count_alexandria
    #WRITING TIME TICKETS
    sheet['B32'] = round(((time_ecc/comp_count_ecc) if comp_count_ecc !=0 else 0),2)
    sheet['B33'] = round(((time_filenet/comp_count_filenet) if comp_count_filenet !=0 else 0),2)
    sheet['B34'] = round(((time_seds/comp_count_seds) if comp_count_seds !=0 else 0),2)
    sheet['B35'] = round(((time_kofax/comp_count_kofax) if comp_count_kofax !=0 else 0),2)
    sheet['B36'] = round(((time_ilinx/comp_count_ilinx) if comp_count_ilinx !=0 else 0),2)
    sheet['B37'] = round(((time_alexandria/comp_count_alexandria) if comp_count_alexandria !=0 else 0),2)
    xfile.save(src_filename)

def ground_zero_wo(src,dir_name):
    import logging,datetime
    LOG_FILENAME = 'SCRIPT.log'
    xl_sheet = gfd.get_sheet_readable(current_dir=dir_name, filename=src, sheet_index=1)
    generate_column_indices(xl_sheet)
    calculations(xl_sheet)
    #display()
    write_file(src)
    file_dest_path = join(dirname(dirname(abspath(__file__))), dir_name, src)
    print("Work Order Data written successfully in "+file_dest_path)
    logging.basicConfig(filename=LOG_FILENAME,level=logging.INFO)
    logging.info("\n Work Order Data written SUCCESSFULLY - "+file_dest_path+" : "+str(datetime.datetime.now()))
