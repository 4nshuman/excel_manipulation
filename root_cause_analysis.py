from __future__ import print_function
from os.path import join, dirname, abspath
import openpyxl, xlrd, get_file_data as gfd

def generate_column_indices(cal_sheet):
    global rc_col_idx, app_col_idx
    # PULLING UP RECORD COLUMNS
    num_cols = cal_sheet.ncols   # Number of columns
    for col_idx in range(0, num_cols):  # Iterate through columns
        cell_obj = cal_sheet.cell(0, col_idx)  # Get cell object by row, col
        #print(cell_obj.value)
        if(cell_obj.value == 'Root Cause'):
            rc_col_idx=col_idx
        if(cell_obj.value == 'Application Name'):
            app_col_idx=col_idx

def calculations(cal_sheet):
    global rca_ecc, rca_seds, rca_filenet, rca_kofax, rca_ilinx, rca_alexandria, consolidated_dict
    # GETTING RC COUNT
    consolidated_dict = {}
    rca_ecc = {}
    rca_seds = {}
    rca_filenet = {}
    rca_kofax = {}
    rca_ilinx = {}
    rca_alexandria = {}
    for row_idx in range(1, cal_sheet.nrows):    # Iterate through rows
        cell_obj = cal_sheet.cell(row_idx, rc_col_idx)  # Get cell object by row, col
        if(cell_obj.value==''):
            cell_obj.value='OTHERS'
        if(cal_sheet.cell(row_idx, app_col_idx).value=='ECC'):
            if(not cell_obj.value in rca_ecc):
                rca_ecc[cell_obj.value]=1
            else:
                rca_ecc[cell_obj.value]+=1
        elif(cal_sheet.cell(row_idx, app_col_idx).value=='SEDS'):
            if(not cell_obj.value in rca_seds):
                rca_seds[cell_obj.value]=1
            else:
                rca_seds[cell_obj.value]+=1
        elif(cal_sheet.cell(row_idx, app_col_idx).value=='FileNet'):
            if(not cell_obj.value in rca_filenet):
                rca_filenet[cell_obj.value]=1
            else:
                rca_filenet[cell_obj.value]+=1
        elif(cal_sheet.cell(row_idx, app_col_idx).value=='Kofax'):
            if(not cell_obj.value in rca_kofax):
                rca_kofax[cell_obj.value]=1
            else:
                rca_kofax[cell_obj.value]+=1
        elif(cal_sheet.cell(row_idx, app_col_idx).value=='ILINX'):
            if(not cell_obj.value in rca_ilinx):
                rca_ilinx[cell_obj.value]=1
            else:
                rca_ilinx[cell_obj.value]+=1
        elif(cal_sheet.cell(row_idx, app_col_idx).value=='Alexandria'):
            if(not cell_obj.value in rca_alexandria):
                rca_alexandria[cell_obj.value]=1
            else:
                rca_alexandria[cell_obj.value]+=1
    consolidated_dict['rca_ecc'] = rca_ecc
    consolidated_dict['rca_seds'] = rca_seds
    consolidated_dict['rca_filenet'] = rca_filenet
    consolidated_dict['rca_kofax'] = rca_kofax
    consolidated_dict['rca_ilinx'] = rca_ilinx
    consolidated_dict['rca_alexandria'] = rca_alexandria

def display():
    for key in consolidated_dict['rca_seds'].keys():
        print (key+" : "+str(consolidated_dict['rca_seds'][key])+"\n")

def get_sheet(sheet_name):
    global xfile
    xfile = openpyxl.load_workbook(src_file,  read_only=False, keep_vba=True)
    return (xfile.get_sheet_by_name(sheet_name))

def write_rca(readable_sheet, write_sheet_name):

    writing_sheet = get_sheet(write_sheet_name)

    num_cols = readable_sheet.ncols   # Number of columns
    for row_idx in range(0,readable_sheet.nrows):
        for col_idx in range(0, num_cols):  # Iterate through columns
            cell_obj = readable_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
            #print (cell_obj.value)
            if(cell_obj.value == 'Root Cause'):
                rc_col_idx=col_idx+1
                rc_row_idx=row_idx+1

    row_increment=1
    for key in consolidated_dict[write_sheet_name.lower().replace('-','_')].keys():
        writing_sheet.cell(row=rc_row_idx+row_increment, column=rc_col_idx).value = key
        writing_sheet.cell(row=rc_row_idx+row_increment, column=rc_col_idx+1).value = consolidated_dict[write_sheet_name.lower().replace('-','_')][key]
        row_increment+=1
    xfile.save(src_file)
    global msg
    msg="Root Cause Analysis for sheet : "+write_sheet_name+" filled successfully"
    print(msg)
    return msg

def ground_zero_rc(src_f,dir_name):
    global src_file
    src_file = src_f
    sheet = gfd.get_sheet_readable(current_dir=dir_name, filename=src_f, sheet_index=3)
    generate_column_indices(sheet)
    calculations(sheet)

    import logging,datetime
    LOG_FILENAME = 'SCRIPT.log'
    logging.basicConfig(filename=LOG_FILENAME,level=logging.INFO)
    file_dest_path = join(dirname(dirname(abspath(__file__))), dir_name, src_f)

    write_rca(gfd.get_sheet_readable(src_f, 20, dir_name), 'RCA-ECC')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    write_rca(gfd.get_sheet_readable(src_f, 20, dir_name), 'RCA-FileNet')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    write_rca(gfd.get_sheet_readable(src_f, 20, dir_name), 'RCA-SEDS')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    write_rca(gfd.get_sheet_readable(src_f, 20, dir_name), 'RCA-Kofax')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    write_rca(gfd.get_sheet_readable(src_f, 20, dir_name), 'RCA-ILINX')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    write_rca(gfd.get_sheet_readable(src_f, 20, dir_name), 'RCA-Alexandria')
    logging.info("\n"+msg+file_dest_path+" : "+str(datetime.datetime.now()))
    #display()
